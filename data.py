from openpyxl import load_workbook

data_file='Book1.xlsx'

wb=load_workbook(data_file)
ws= wb['Sheet1']
all_rows=list(ws.rows)
book_data=[]
for row in all_rows[1:]:
    name=row[0].value
    isbn=row[1].value
    author=row[2].value
    dict={
        'name':name,
        'isbn':isbn,
        'author':author
        }
    book_data.append(dict)

data_file='Book_populate.xlsx'
wb=load_workbook(data_file)
ws= wb['Sheet1']
all_rows=list(ws.rows)
book_data_populate=[]
for row in all_rows[1:]:
    name=row[0].value
    isbn=row[1].value
    author=row[2].value
    dict={
        'name':name,
        'isbn':isbn,
        'author':author
        }
    book_data_populate.append(dict)

